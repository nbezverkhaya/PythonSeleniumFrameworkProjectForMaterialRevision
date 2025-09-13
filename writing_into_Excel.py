# working with Excel tables
import openpyxl
# working with CSV
# import pandas
book = openpyxl.load_workbook("/Users/natulik/PycharmProjects/python_testing/PythonSeleniumFrameworkProjectForMaterialRevision/python_demo.xlsx")
sheet = book.active

# cell = sheet.cell(4, 1)
# print(cell.value)
#
sheet.cell(4,2).value = "Nataliia"
# print(sheet.cell(4,2).value)
# print(sheet['B4'].value)
#
# print(sheet.max_row)
# print(sheet.max_column)

for i in range (1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range (1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)